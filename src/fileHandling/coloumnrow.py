import requests
from openpyxl import Workbook

wb=Workbook()
filepath = 'C:/JavaProgramming/ConsumingAPI/resources/excel/product_data.xlsx'
sheet = wb.create_sheet("newsheet")

k=1
for i in range(4):
    sheet.cell(row=k,column=1).value="hello"
    sheet.cell(row=k,column=2).value="how"
    sheet.cell(row=k,column=3).value="are"
    sheet.cell(row=k,column=4).value="you"
    k=k+1
wb.save(filepath)
