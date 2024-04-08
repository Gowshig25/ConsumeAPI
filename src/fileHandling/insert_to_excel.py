import openpyxl
from openpyxl.workbook import Workbook


def write_in_excel(data):
    wb= Workbook()
    sheet = wb.active
    sheet.append(["Product ID", "Product Name", "Product Quantity", "Product Price"])

    '''  for i, product in enumerate(data, start=2):
    # Iterate over keys and values of each product
        for j, (key, value) in enumerate(product.items(), start=1):
        # Write the value to the corresponding cell
            sheet.cell(row=i, column=j).value = value'''
    i=1
    for product in data:
        i=i+1
        j=1
        for (key,value) in product.items():
            sheet.cell(row=i, column=j).value = value
            j+=1
    filepath = 'C:/JavaProgramming/ConsumingAPI/resources/excel/product_data.xlsx'
    wb.save(filepath)
    print('the data is added in excel successfully')
